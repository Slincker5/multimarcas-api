import json
import os
import requests
from openpyxl import load_workbook
from datetime import datetime
import uuid

def main():
    if len(sys.argv) != 4:
        sys.exit(f"Usage: {sys.argv[0]} <URL> <TOKEN> <UUID>")

    url, token, user_uuid = sys.argv[1], sys.argv[2], sys.argv[3]

    headers = {'Authorization': 'Bearer ' + token}
    response = requests.get(url, headers=headers)

    if response.status_code != 200:
        sys.exit(f"Error in server response: {response.status_code}")

    try:
        resultados = response.json()
    except json.JSONDecodeError as e:
        sys.exit(f"Error decoding JSON: {e}")

    wb = load_workbook("public/documentos/PLANTILLA_CINTILLOS.xlsx")
    ws = wb["DATA"]

    for i, resultado in enumerate(resultados):
        row = i + 3
        col_prefix = "B"
        if i >= 226:
            row = (i - 226) + 3
            col_prefix = "E"
        ws[f"{col_prefix}{row}"] = resultado['barra']
        ws[f"{chr(ord(col_prefix) + 1)}{row}"] = resultado['descripcion']
        ws[f"{chr(ord(col_prefix) + 2)}{row}"] = "$" + resultado['precio']

    parent_dir_path = os.path.join("public/cintillos", user_uuid)
    os.makedirs(parent_dir_path, exist_ok=True)

    path_uuid = str(uuid.uuid4())
    child_dir_path = os.path.join(parent_dir_path, path_uuid)
    os.makedirs(child_dir_path, exist_ok=True)

    now = datetime.now().strftime("%Y-%m-%d-%H%M%S")
    file_name = f"CINTILLOS-{now}.xlsx"
    file_path = os.path.join(child_dir_path, file_name)

    wb.save(file_path)

    tmp_dir = os.path.join(os.path.dirname(file_path), "tmp")
    os.makedirs(tmp_dir, exist_ok=True)
    tmp_file_path = os.path.join(tmp_dir, os.path.basename(file_path))

    response = {
        "status": "OK",
        "message": "Archivo generado con éxito",
        "path_name": file_name,
        "path_complete": file_path,
        "path_tmp": tmp_dir,
        "path_tmp_full": tmp_file_path,
        "path_uuid": path_uuid,
        "user_uuid": user_uuid,
        "cantidad": len(resultados)
    }

    print(json.dumps(response))

if __name__ == "__main__":
    import sys
    main()
